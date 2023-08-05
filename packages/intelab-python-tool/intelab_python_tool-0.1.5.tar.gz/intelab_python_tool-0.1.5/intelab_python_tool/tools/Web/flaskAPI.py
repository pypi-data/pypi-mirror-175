from datetime import datetime
import os
from intelab_python_tool.tools.log import log, log_init
from flask import Flask, jsonify, request, make_response, json
from flask.views import MethodView
from apscheduler.schedulers.blocking import BlockingScheduler
import requests
import openpyxl
import threading
from gevent import pywsgi
from dynaconf import settings

"""
使用方法如下：
    
    from intelab_python_tool.tools.Web.flaskAPI import Monitor
    Mt = Monitor(scheduler_list)
    Mt.run()
    
    scheduler_list: list
    such as :  [
                [service_name, monitor_log_path, monitor_type, scheduler_period_type, 
                scheduler_period, dingtalk_function
                restart_flag, args],
                实例：
                ['event_verifier', '/var/log/utilization_producer/log/timing_utilization_producer.info.log.',
                'l', 'm', 3, 0, 1],
                ]
    
    service_name: 如果监控日志，则是重启服务时的服务名；如果是接口测试，则是url
    monitor_log_path: 如果监控日志，则是监控的日志文件的绝对路径；如果是接口测试，则是访问接口的方式get or post
    monitor_type: 'l'表示监控日志, 'a'表示监控接口
    scheduler_period_type: 定时任务的类型，'h'表示按小时，'m'表示按分钟
    scheduler_period:  定时任务的时长， 多少小时/多少分钟
    dingtalk_function: 发送报警的函数 为0表示无效
    restart_flag: 是否需要重启服务
    args: 接口测试时需要传入的参数
    
"""

TIMEZONE = 'Asia/Shanghai'


class Monitor:

    def __init__(self, scheduler_list=None, port=5000):
        # 初始化输入参数
        self.scheduler_list = scheduler_list
        self.service_list = self.scheduler_init()
        # 初始化本地文件储存相应信息
        self.create_local_excel()

        self.time = datetime.now().strftime("%Y-%m-%d")
        self.hour = datetime.now().hour

        self.port = settings.get('MONITOR_PORT', port)

    def run(self, logs_path='/var/log/event/control_log/', host='0.0.0.0'):
        log_init('health_monitor', debug=False, log_path=logs_path)
        try:
            if isinstance(self.port, str):
                port = eval(self.port)
            elif isinstance(self.port, int):
                port = self.port
            else:
                log.info('健康检测端口配置有误，默认使用5000')
                port = 5000

            t1 = threading.Thread(target=self.start_scheduler)
            t2 = threading.Thread(target=self.api_run, args=[host, port])

            t1.start()
            t2.start()
        except Exception as e:
            log.info(e)


    def api_run(self, host, port):
        app = Flask(__name__)

        @app.route('/api/ai/service_monitor', methods=['GET'])
        def get_service_health():
            record_path_list = []
            for service in self.service_list:
                record_path_list.append(service.name)
            path_fix = '/var/log/event/control_log/'
            error_time_list = []
            for _ in record_path_list:
                record_path = path_fix + _ + '_monitor.xlsx'
                wb = openpyxl.load_workbook(record_path)
                ws = wb.active
                error_times = ws['C1'].value
                error_time_list.append(error_times)
                wb.close()

            if max(error_time_list) < 3:
                response = {'msg': '服务正常', 'code': 200,
                            'data': error_time_list}
                return jsonify(response)
            else:
                msg = {'msg': '服务异常'}
                response = make_response(json.dumps(msg), 400)
                return response

        # app.config['env'] = 'development'
        # app.run(host=host, port=port)

        server = pywsgi.WSGIServer((host, port), app)
        server.serve_forever()

    def monitor_log_run(self, service):

        log.info(f"本次执行健康检测的服务为{service.name}")

        wb = openpyxl.load_workbook(f'/var/log/event/control_log/{service.name}_monitor.xlsx')
        ws = wb.active
        old_data_bytes = ws['A1'].value
        check_times = ws['B1'].value
        error_times = ws['C1'].value

        try:
            new_data_bytes = os.stat(f"{service.log_path}" + f"{self.time}").st_size
        except Exception as e:
            log.info("发生未知错误，无法获取文件大小，请检查文件是否存在")
            new_data_bytes = 0

        log.info('主监控日志文件变化前大小为{}，变化后大小为{}'.format(old_data_bytes, new_data_bytes))

        if service.restart_type:

            if new_data_bytes != old_data_bytes and error_times == 0:
                check_times = 0
            else:
                check_times += 1

            log.info(f'主监控日志check_times为{check_times}')
            if check_times > 3:
                if error_times == 0:
                    log.info(f'check_times大于3，重启服务')
                    service_name = service.name if not service.restart_name else service.restart_name
                    self.restart_service(service_name, service.dingtalk_function, service.restart_flag)
                    check_times = 0
                    error_times += 1
                else:
                    error_times = 4
            else:
                error_times = 0

        else:

            if new_data_bytes != old_data_bytes:
                check_times = 0
            else:
                check_times += 1

            if error_times > 3:
                error_times = 0

            log.info(f'主监控日志check_times为{check_times}')
            if check_times > 3:
                log.info(f'check_times大于3，重启服务')
                service_name = service.name if not service.restart_name else service.restart_name
                self.restart_service(service_name, service.dingtalk_function, service.restart_flag)
                check_times = 0
                error_times += 1

        hour_str = f"{self.hour}" if self.hour > 9 else f"0{self.hour}"
        log_check = os.system(f"cat {service.log_path}" + f"{self.time} | grep {self.time}T{hour_str}")
        log.info(f"执行shell: cat {service.log_path}" + f"{self.time} | grep {self.time}T{hour_str}")

        if log_check != 0:
            log.info(f"{self.time}T{hour_str} 无法在日志文件中查询到，请及时查看处理")

        ws['A1'].value = new_data_bytes
        ws['B1'].value = check_times
        ws['C1'].value = error_times
        wb.save(f'/var/log/event/control_log/{service.name}_monitor.xlsx')

        self.time = datetime.now().strftime("%Y-%m-%d")
        self.hour = datetime.now().hour

    def restart_service(self, service_name, ding_talk_function, restart_flag) -> None:
        if not restart_flag:
            return
        log.info(f"正在重启{service_name}服务, supervisorctl restart {service_name}")
        res = os.system(f"supervisorctl restart {service_name}")
        log.info(res)
        if res == 0:
            log.info(f"重启{service_name}完成")
        else:
            self.send_alarm2developer(f"重启{service_name}失败，请开发人员介入查看", ding_talk_function=ding_talk_function)

    @staticmethod
    def send_alarm2developer(msg: str, ding_talk_function) -> None:
        if isinstance(ding_talk_function, int):
            return
        log.info('警告已发送至开发者')
        ding_talk_function(msg)

    def create_local_excel(self) -> None:
        for service in self.service_list:
            if service.monitor_type == 'l':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = 0    # 储存监控日志大小
                ws['B1'] = 0    # 储存check_times
                ws['C1'] = 0    # 储存error_times
                wb.save(f'/var/log/event/control_log/{service.name}_monitor.xlsx')
                wb.close()

    def test_request(self, service):
        url = service.name
        request_type = service.log_path
        data = service.args

        if request_type == 'get':
            res = requests.get(url) if not data else requests.get(url, data=data)
        else:
            res = requests.post(url) if not data else requests.post(url, data=data)

        if res.status_code == 200:
            return True
        else:
            msg = f'{service.name}接口调用失败，请及时查看'
            self.send_alarm2developer(msg, service.dingtalk_function)

    def start_scheduler(self) -> None:

        scheduler = BlockingScheduler(
            {'apscheduler.timezone': TIMEZONE}
        )
        for service in self.service_list:
            if service.monitor_type == 'l':
                if service.scheduler_period_type == 'h':
                    scheduler.add_job(self.monitor_log_run, 'interval', hours=service.scheduler_period, args=[service])
                else:
                    scheduler.add_job(self.monitor_log_run, 'interval', minutes=service.scheduler_period, args=[service])
            elif service.monitor_type == 'a':
                if service.scheduler_period_type == 'h':
                    scheduler.add_job(self.test_request, 'interval', hours=service.scheduler_period, args=[service])
                else:
                    scheduler.add_job(self.test_request, 'interval', minutes=service.scheduler_period, args=[service])
            else:
                log.info(f"{service.name}服务监控类型不存在，服务监控类型目前仅支持l,a两种")
        scheduler.start()

    def scheduler_init(self):
        if self.scheduler_list:
            return [Service(_) for _ in self.scheduler_list]
        else:
            config_scheduler_list = settings.get('MONITOR_SCHEDULERS')
            if isinstance(config_scheduler_list, str):
                config_scheduler_list = eval(config_scheduler_list)
            res = []
            for _ in config_scheduler_list:
                if isinstance(_, list) or isinstance(_, dict):
                    res.append(Service(_))
                else:
                    log.info(f'{_}的配置类型不是list或dict，请检查')
            return res


class ServiceHealthApi(MethodView):

    def get(self, record_paths: str):
        record_path_list = eval(record_paths)
        path_fix = '/var/log/event/control_log/'
        error_time_list = []
        for _ in record_path_list:
            record_path = path_fix + _ + '_monitor.xlsx'
            wb = openpyxl.load_workbook(record_path)
            ws = wb.active
            error_times = ws['C1'].value
            error_time_list.append(error_times)
            wb.close()

        if max(error_time_list) >= 3:
            response = {'msg': '算法服务重启三次仍无法正常运行，请相关运维人员检查服务配置或网络', 'code': 400,
                        'data': error_time_list}
        else:
            response = {'msg': '服务正常', 'code': 200, 'data': error_time_list}
        return jsonify(response)


class Service:

    def __init__(self, msg):
        if isinstance(msg, list):
            self.name = msg[0]
            self.log_path = msg[1]
            self.monitor_type = msg[2]
            self.scheduler_period_type = msg[3]
            self.scheduler_period = msg[4]
            self.dingtalk_function = msg[5]
            self.restart_flag = msg[6]
            if len(msg) == 8:
                self.restart_name = msg[7]
            elif len(msg) < 8:
                self.restart_name = 0
            if len(msg) == 9:
                self.args = msg[8] if msg[8] else None
        elif isinstance(msg, dict):
            self.name = msg['service_name']
            self.log_path = msg['log_path']
            self.monitor_type = msg['monitor_type']
            self.scheduler_period_type = msg['scheduler_period_type']
            self.scheduler_period = msg['scheduler_period']
            self.dingtalk_function = msg['ding_talk_function']
            self.restart_flag = msg['restart_flag']
            self.restart_type = msg.get('restart_type', 0)
            self.restart_name = msg.get('restart_name', 0)
            self.args = msg.get('args', None)


if __name__ == '__main__':
    a = '[{"service_name": "intelab-video-cron", "log_path": "/var/log/intelab-video/logs/cron/cron.info.log.", "monitor_type": "l", "scheduler_period_type": "m","scheduler_period": 20,"ding_talk_function": 0,"restart_flag": 1 }, {"service_name": "intelab-video-record", "log_path": "/var/log/intelab-video/logs/record/record.info.log.", "monitor_type": "l", "scheduler_period_type": "m","scheduler_period": 20,"ding_talk_function": 0,"restart_flag": 1 }]'
    b = eval(a)
    res = []
    for _ in b:
        if isinstance(_, list) or isinstance(_, dict):
            res.append(Service(_))
        else:
            log.info(f'{_}的配置类型不是list或dict，请检查')
    s = res
    c =s


