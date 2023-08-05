import pandas as pd
import csv
import os
from openpyxl.reader.excel import ExcelReader

# 须知：excel和csv对应在office类软件中的xlsx, xls表格文件


class Excel(ExcelReader):

    def __init__(self, path, sheet_index=0, sheet_name=None):
        """
        打开一个xlsx文件，并根据下标或名称选取sheet，进行相关初始化
        使用方法如下：

        self.worksheet 获取sheet, 为了节省代码, self.ws = self.worksheet
        >> 以下名称已经说明了用法，不再介绍
        self.ws.title
        self.ws.max_row
        self.ws.min_row
        self.ws.max_column
        self.ws.min_column
        self.ws.rows()
        self.ws.columns()
        self.ws.values()
        self.ws.iter_rows()
        self.ws.iter_columns()
        self.ws.append() >> 参数是iter类型(如list)的数据，按行输入, 如果想输入其它非iter类型的数据，可以考虑先本库的datatype转换数据类型
        >> from intelab_python_tool.tools.datatype import ...
        self.ws[] >> 该方法中括号中可以输入 '英文字母'、'数字'、'单元格名称' 来分别获取列对象、行对象、单元格的值

        :param path:
        :param sheet_index:
        :param sheet_name:
        """
        super(Excel, self).__init__(path, read_only=False, keep_vba=False, data_only=False, keep_links=True)
        self.read()

        self.sheet_names = self.wb.sheetnames
        self.encoding = self.wb.encoding
        self.properties = self.wb.properties
        self.remove = self.wb.remove
        self.create_excel = self.wb.create_sheet

        if not sheet_name:
            self.work_sheet = self.wb.worksheets[sheet_index]
        else:
            self.work_sheet = self.wb[sheet_name]

        self.ws = self.work_sheet


def get_excel_one_column(local_path, column_name: str) -> list:
    """
    获取 excel 其中一列并作为列表输出
    :param local_path:
    :param column_name: 列名
    :return:
    """
    df = pd.read_excel(local_path)
    column_list = df[column_name].values.tolist()
    return column_list


def csv_write_rows(local_path, data: list) -> None:
    """
    按行写入数据至 csv 文件中, csv 不存在则创建
    """
    f = open(local_path, "a", newline="", encoding='utf-8')
    writer = csv.writer(f)
    writer.writerows(data)
    f.close()


def all_excel2one(local_path, folders_path, sheet_name=0) -> list:
    """
    将文件夹内所有 excel 文件第 sheet_name 张表（默认为0）中的数据提取放入一个新的csv中
    :param local_path:
    :param folders_path:
    :param sheet_name:
    :return:
    """
    row_list = []
    for root, dirs, files in os.walk(folders_path):
        for file in files:
            try:
                read_data = pd.read_excel((folders_path + '/' + file), sheet_name=sheet_name)
                file_data = read_data.values.tolist()
                for row in file_data:
                    row_list.append(row)
            except Exception as e:
                print(e, type(e))
                continue
    csv_write_rows(local_path, row_list)
    return row_list


def xls2xlsm(local_path) -> any:
    """
    将 csv 文件转化为 excel 文件
    :param local_path:
    :return:
    """
    df = pd.read_excel(local_path)
    xlsm_excel = df.to_excel(local_path + 'm', index=False)
    return xlsm_excel


def excel2csv(local_path) -> any:
    """
    将 excel 文件转化为 csv 文件
    :param local_path:
    :return:
    """
    df = pd.read_excel(local_path)
    csv_data = df.to_csv(local_path[:-2], index=False)
    return csv_data

