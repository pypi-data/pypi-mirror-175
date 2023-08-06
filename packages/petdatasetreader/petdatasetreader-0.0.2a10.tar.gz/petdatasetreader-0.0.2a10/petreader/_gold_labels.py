import openpyxl
from petreader.labels import *
from petreader.ProcessInformation import LABEL
from petreader.ProcessInformation import ProcessInformation

pi = ProcessInformation()
graph_folder = './'
def get_examples_data_DFG_GoldStandardLabels(doc_name):
    text = pi.tc.GetDocumentText(doc_name)
    gold_activity_labels = _collect_GoldStandardLabels(doc_name)
    dfg = pi.GetDFG(doc_name, graph_folder)

    #  get edges
    edges = list()
    for edge in dfg.edges:
        source, target = edge
        label_source = gold_activity_labels[source][LABEL]
        label_target = gold_activity_labels[target][LABEL]
        edge = f'{label_source} {RELATION_SEP} {label_target}'
        edges.append(edge)
    return text, edges

def _collect_GoldStandardLabels(doc_name):
    wb = openpyxl.load_workbook('process_data.xlsx')
    ws1 = wb[doc_name]

    column_n_sent = 1
    column_w_index = 2
    column_activity_labels = 6
    column_participant_labels = 9

    #  look for "GOLD ACTIVITY LABEL" in column 6
    gold_activity_labels = dict()
    for row in range(2,ws1.max_row+1):
        if ws1.cell(row, column_activity_labels).value == "GOLD ACTIVITY LABEL":
            row += 3
            while True:
                if not ws1.cell(row, column_activity_labels).value:
                    return gold_activity_labels

                n_sent = ws1.cell(row, column_n_sent).value
                w_index = ws1.cell(row, column_w_index).value

                label = ws1.cell(row, column_activity_labels).value
                label = label.strip()
                performer = ws1.cell(row, column_participant_labels).value
                performer = performer.strip()

                node = (int(n_sent), int(w_index))
                gold_activity_labels[node] = {LABEL: label.lower(),
                                              ACTOR_PERFORMER: performer.lower()}
                row += 1
doc_name='doc-3.2'
print(_collect_GoldStandardLabels(doc_name))