#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook


class ExcelLoader():

    def __init__(self, file, extra_data=None):
        self.excel = load_workbook(file)
        self.extra_data = extra_data
        self.data_num = -1
        self._init()

    def _init(self):
        pass

    def _load_header(self, sheet):
        header = [cell.value for cell in sheet[1] if cell.value]
        return header

    def data(self, sheet_index=None, read_mode="dict", skip_line=0):
        if sheet_index is not None:
            sheets = [self._get_sheet(sheet_index)]
        else:
            sheets = self._get_sheets()
        for sheet in sheets:
            header = self._load_header(sheet) if read_mode == "dict" else None
            for data in self._read_data(header, sheet, read_mode, skip_line):
                yield data

    def _get_sheets(self):
        return self.excel.sheets()

    def get_sheet_num(self):
        return len(self.excel.sheetnames)

    def get_sheet_name(self, index):
        return self.excel.sheetnames[index]

    def _get_sheet(self, index):
        return self.excel.worksheets[index]

    def _read_data(self, header, sheet, read_mode="dict", skip_line=0):
        self.data_num = -1
        for num, row in  enumerate(sheet.iter_rows(skip_line+2)):
            row_data = [self._handle_cell_value(cell) for cell in row]
            if not self._is_empty_row(row_data):
                self.data_num = self.data_num + 1
                if read_mode == "dict":
                    yield self.data_num, self._handle_data(dict(zip(header, row_data[:len(header)])))
                elif read_mode == "list":
                    yield self.data_num, row_data
                else:
                    pass
            else:
                break

    def _handle_cell_value(self, cell):
        value = None
        try:
            if isinstance(cell.value, long):
                value = int(cell.value)
            else:
                value = cell.value
        except Exception:
            value = cell.value
        return value

    def _is_empty_row(self, row_data):
        return not bool([value for value in row_data if value])

    def _handle_data(self, data):
        return data
