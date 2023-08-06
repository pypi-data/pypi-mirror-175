#!/usr/bin/env python3
# -*- coding:utf-8 -*-
'''
# File: excel.py
# Project: Sloppy
# Created Date: 2022-07-18, 09:17:13
# Author: Chungman Kim(h2noda@unipark.kr)
# Last Modified: Wed Oct 19 2022
# Modified By: Chungman Kim
# Copyright (c) 2022 Unipark
# HISTORY:
# Date      	By	Comments
# ----------	---	----------------------------------------------------------
'''
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font, numbers
from openpyxl import Workbook


class Xlsx:
    def __init__(self, filename, sheetname="sheet", idx=0, mode="n"):
        """Mode = n : New Excel File
                  r : Read Excel File

        Args:
            filename (String): Excel File Name
            sheetname (String): Excel Sheet Name
            mode (str, optional): Excel file Mode, Defaults to "n".
        """
        if (mode == "n"):
            self.filename = filename
            self.sheetname = sheetname
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = sheetname
        elif (mode == "r"):
            self.filename = filename
            self.wsname = sheetname
            self.wb = openpyxl.load_workbook(filename)
            ws_name = self.wb.get_sheet_names()
            if sheetname in ws_name:
                self.ws = self.wb[ws_name[0]]
            elif sheetname == "":
                self.ws = self.wb[ws_name[idx]]
            else:
                self.ws = self.wb[sheetname]

    def setting_title(self, title):
        self.ws.title = title

    def header(self, col, row, width, color, val):
        loc = col + row
        if (width != 0):
            self.ws.column_dimensions[col].width = width
        self.self.ws[loc].alignment = Alignment(horizontal="center")
        self.self.ws[loc].border = Border(
            Side("thin"), Side("thin"), Side("thin"), Side("thin"))
        self.self.ws[loc].fill = PatternFill(
            fill_type="solid", start_color=color, end_color=color)
        self.self.ws[loc].font = Font(size="10")
        self.self.ws[loc] = val

    def footer(self, col, row, color, aligment, val):
        loc = col + row
        self.self.ws[loc].alignment = Alignment(horizontal=aligment)
        self.self.ws[loc].border = Border(Side("thin"), Side(
            "thin"), Side("thin"), Side("thin"))
        self.self.ws[loc].fill = PatternFill(
            fill_type="solid", start_color=color, end_color=color)
        self.self.ws[loc].font = Font(size="10")
        self.self.ws[loc] = val

    def datacell(self, col, row, aligment, val):
        loc = col + row
        self.self.ws[loc].alignment = Alignment(horizontal=aligment)
        self.self.ws[loc].border = Border(
            Side("thin"), Side("thin"), Side("thin"), Side("thin"))
        self.self.ws[loc].font = Font(size="10")
        self.self.ws[loc] = val

    def cell_style(self, col, row, alignment="center", border="thin", fill="solid", color=None, font="D2Coding", fontsize="10"):
        """Color = #000000, #FF0000

        Args:
            col (_type_): _description_
            row (_type_): _description_
            alignment (str, optional): _description_. Defaults to "center".
            border (str, optional): _description_. Defaults to "thin".
            fill (str, optional): _description_. Defaults to "solid".
            color (str, optional): _description_. Defaults to "".
            font (str, optional): _description_. Defaults to "D2Coding".
            fontsize (str, optional): _description_. Defaults to "10".
        """
        loc = col + row
        side = Side(border)
        self.self.ws[loc].alignment = Alignment(horizontal="center")
        self.self.ws[loc].border = Border(side, side, side, side)
        self.self.ws[loc].fill = PatternFill(
            fill_type=fill, start_color=color, end_color=color)
        self.self.ws[loc].font = Font(size=fontsize)

    def cell_displaytype(self, col, row, displaytype):
        if (displaytype == "Currency"):
            self.ws.cell(
                row=row, column=col).number_format = numbers.BUILTIN_FORMATS[3]

    def excel_load(self, filename, sheetname):
        self.filename = filename
        self.wb = openpyxl.load_workbook(filename)
        self.self.ws = self.wb[sheetname]

    def excel_save(self):
        self.wb.save(self.filename)

    def select_sheet(self, sheetname):
        self.self.ws = self.wb[sheetname]

    def create_sheet(self, sheetname, loc=None):
        """loc = 0  : First
                 -1 : Second from the back
                 1  : Second
        Args:
            sheetname (_type_): _description_
            loc (_type_): _description_
        """
        self.self.ws = self.wb.create_sheet(sheetname, loc)

    def rename_sheetname(self, b_sheetname, a_sheetname):
        self.wb[b_sheetname].title = a_sheetname
