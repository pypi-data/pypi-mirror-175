#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2022-06-16 10:07
# @Site    :
# @File    : dealFile.py
# @Software: PyCharm
import os

import openpyxl as op

def create_excel_xlsx(path, sheet_name):
    workbook = op.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)

def write_excel_xlsx_append(path,name, value, truncate_sheet=False):
    """
    :param path: ./demo/test.xlsx
    :param value: [[],[]]
    :param truncate_sheet:
    :return:
    """
    # 如果不存在就创建该excel
    if not os.path.exists(f"{path}/{name}", ):
        os.makedirs(path)
        create_excel_xlsx(f"{path}/{name}", 'Sheet1')
    data = op.load_workbook(f"{path}/{name}", )
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if(truncate_sheet): #truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        # print(sheet.title)  # 输出表名
        startrows = sheet.max_row  # 获得行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=startrows + i + 1, column=j + 1, value=str(value[i][j]))
    data.save(f"{path}/{name}",)
    print("xlsx格式表格追加写入数据成功！")

def read_excel_xlsx(path, sheet_name):
    workbook = op.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def read_excel(path: str, sheet: str):
    """读取excel"""
    # 打开工作簿
    workbook = op.load_workbook(path)
    # 获取表单
    sheet = workbook[sheet]
    # 获取最大行数
    max_row = sheet.max_row
    # 获取最大列数
    max_column = sheet.max_column
    lr = tuple([tuple([sheet.cell(row=row, column=column).value for column in range(1, max_column+1)]) for row in range(1, max_row+1)])
    print(lr)
    return lr


if __name__ == '__main__':
    path = "./test"
    value = [[1,1,2,3,3,4],[5,6,3,12,5]]
    write_excel_xlsx_append(path,"test.xlsx",value)
